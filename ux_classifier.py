import os
import json
import pandas as pd
from dotenv import load_dotenv
from openai import OpenAI
from rapidfuzz import fuzz
from utils import get_cache_file_path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import CellIsRule, FormulaRule

# ========== SETUP ==========
load_dotenv(dotenv_path=".env")
api_key = os.getenv("OPENAI_API_KEY")
if not api_key:
    raise ValueError("OpenAI API key not found in .env")

client = OpenAI(api_key=api_key)

INPUT_FILE = "data/small_test.xlsx"    # Must have "Comment" and "Tag" columns
OUTPUT_FILE = "data/ux_test_output.xlsx"
CACHE_DIR = "ux_cache"
os.makedirs(CACHE_DIR, exist_ok=True)

# ========== CONFIG ==========
UX_KEYWORDS = [
    "hard to find", "confusing", "can't find", "unclear",
    "too many steps", "not intuitive", "difficult to", "didn't work", "didn't work",
    "site is slow", "popup", "didn't expect", "unexpected", "everything keeps changing", "broken",
    "interrupt", "interrupted", "sucks", "impossible", "no longer works", "not available",
    ]

UX_PROMPT = """Does the following user comment raise a UX (user experience) concern? If yes, classify the type of concern.

UX Types:
- Navigation
- Clarity/Language
- Technical/Performance
- Interruption/Flow
- Other

Comment: "{comment}"

Output:
UX Concern: Yes/No
UX Type: [Type if Yes, else None]
"""

SENTIMENT_PROMPT = """You are a sentiment classifier. For each comment, classify it as either Positive or Negative.

Only use the two labels: Positive or Negative.
Neutral, unclear, or mixed sentiments should be treated as Negative.

Examples:
Comment: "I love this new layout. It's easy to use and fast."
Sentiment: Positive

Comment: "I couldn't find what I was looking for. Very frustrating."
Sentiment: Negative

Comment: "The site worked fine but then crashed while I was checking out."
Sentiment: Negative

Comment: "{comment}"

Output:
Sentiment:"""

# ========== HELPERS ==========

def extract_sentiment(output):
    for line in output.splitlines():
        lower_line = line.lower()
        if "negative" in lower_line:
            return "Negative"
        elif "positive" in lower_line:
            return "Positive"
    return "Unknown"

def contains_ux_keywords_fuzzy(comment, keywords=UX_KEYWORDS, threshold=80):
    comment_lower = comment.lower()
    return any(fuzz.partial_ratio(comment_lower, k.lower()) >= threshold for k in keywords)

def analyze_sentiment(comment):
    prompt = SENTIMENT_PROMPT.format(comment=comment)
    model = "gpt-3.5-turbo"
    task = "sentiment"
    cache_path = get_cache_file_path(comment, task, model, prompt)

    if os.path.exists(cache_path):
        with open(cache_path, "r") as f:
            return json.load(f)["sentiment"]

    try:
        response = client.chat.completions.create(
            model=model,
            messages=[{"role": "user", "content": prompt}],
            temperature=0.0
        )
        output = response.choices[0].message.content.strip()
        sentiment = extract_sentiment(output)
        with open(cache_path, "w") as f:
            json.dump({"sentiment": sentiment}, f)
        return sentiment
    except Exception as e:
        print(f"[Sentiment Error] {e}")
        return "Error"

def analyze_ux_concern(comment):
    prompt = UX_PROMPT.format(comment=comment)
    model = "gpt-3.5-turbo"
    task = "ux"
    cache_path = get_cache_file_path(comment, task, model, prompt)

    if os.path.exists(cache_path):
        with open(cache_path, "r") as f:
            cached = json.load(f)
            return cached["ux_concern"], cached["ux_type"]

    try:
        response = client.chat.completions.create(
            model=model,
            messages=[{"role": "user", "content": prompt}],
            temperature=0.0
        )
        output = response.choices[0].message.content.strip()
        concern, concern_type = "No", "None"
        for line in output.split("\n"):
            if line.lower().startswith("ux concern:"):
                concern = line.split(":", 1)[1].strip()
            elif line.lower().startswith("ux type:"):
                concern_type = line.split(":", 1)[1].strip()
        with open(cache_path, "w") as f:
            json.dump({"ux_concern": concern, "ux_type": concern_type}, f)
        return concern, concern_type
    except Exception as e:
        print(f"[UX Concern Error] {e}")
        return "Error", "Error"


def is_similar_to_others(comment, all_comments, threshold=57, min_similar=2):
    """
    Check if a comment is semantically similar to at least `min_similar` other comments.
    """
    similar_count = 0
    for other_comment in all_comments:
        if comment != other_comment:
            similarity = fuzz.partial_ratio(comment.lower(), other_comment.lower())
            if similarity >= threshold:
                similar_count += 1
            if similar_count >= min_similar:
                return True
    return False

def compute_ux_score(sentiment, keyword_hit, ux_flag, similarity_flag):
    score = 0
    if sentiment.lower() == "negative":
        score += 1
    if keyword_hit:
        score += 1
    if ux_flag.lower() == "yes":
        score += 1
    if similarity_flag:
        score += 1
    return score  # Max = 4

# ========== MAIN ==========
def main():
    df = pd.read_excel(INPUT_FILE)
    if not {"Comment", "Tag"}.issubset(df.columns):
        raise ValueError("Input file must contain 'Comment' and 'Tag' columns")

    prediction_cols = {
        "Sentiment": [],
        "UX Keyword Flag": [],
        "UX Concern GPT": [],
        "UX Type GPT": [],
        "Similarity Flag": [],
        "UX Composite Score": []
    }

    all_comments = df["Comment"].tolist()

    for _, row in df.iterrows():
        comment = row["Comment"]
        print(f"Processing: {comment[:60]}...")

        keyword_flag = contains_ux_keywords_fuzzy(comment)
        sentiment = analyze_sentiment(comment)
        ux_flag, ux_type = analyze_ux_concern(comment)
        similarity_flag = is_similar_to_others(comment, all_comments)
        score = compute_ux_score(sentiment, keyword_flag, ux_flag, similarity_flag)

        prediction_cols["Sentiment"].append(sentiment)
        prediction_cols["UX Keyword Flag"].append(keyword_flag)
        prediction_cols["UX Concern GPT"].append(ux_flag)
        prediction_cols["UX Type GPT"].append(ux_type)
        
        # Only show similarity flag if sentiment is not positive
        if similarity_flag and sentiment.lower() != "positive":
            prediction_cols["Similarity Flag"].append("FLAG")
        else:
            prediction_cols["Similarity Flag"].append("")
            
        prediction_cols["UX Composite Score"].append(score)

    for col_name, values in prediction_cols.items():
        df[col_name] = values

    # Save to Excel
    df.to_excel(OUTPUT_FILE, index=False)
    
    # Apply conditional formatting
    apply_conditional_formatting(OUTPUT_FILE)
    
    print(f"\n🎉 Done! Output written to {OUTPUT_FILE}")

def apply_conditional_formatting(file_path):
    """Apply conditional formatting to highlight composite scores above 3 in red with white text"""
    try:
        # Load the workbook
        wb = load_workbook(file_path)
        ws = wb.active
        
        # Find the column index for "UX Composite Score"
        header_row = 1  # Assuming headers are in row 1
        score_col = None
        
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=header_row, column=col).value
            if cell_value == "UX Composite Score":
                score_col = col
                break
        
        if score_col is None:
            print("Warning: 'UX Composite Score' column not found. Conditional formatting not applied.")
            return
        
        # Get the column letter
        from openpyxl.utils import get_column_letter
        col_letter = get_column_letter(score_col)
        
        # Define the range for conditional formatting (excluding header)
        cell_range = f"{col_letter}2:{col_letter}{ws.max_row}"
        
        # Create a red fill and white font
        red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        white_font = Font(color="FFFFFFFF")  # White font color
        
        # Add conditional formatting rule using FormulaRule to apply both fill and font
        rule = FormulaRule(
            formula=[f'${col_letter}2>3'],  # Highlight scores above 3
            stopIfTrue=True,
            fill=red_fill,
            font=white_font
        )
        
        ws.conditional_formatting.add(cell_range, rule)
        
        # Save the workbook
        wb.save(file_path)
        print("Conditional formatting applied successfully.")
    except Exception as e:
        print(f"Error applying conditional formatting: {e}")

if __name__ == "__main__":
    main()
